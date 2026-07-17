import csv
import json
import os

from dotenv import load_dotenv
from openpyxl import load_workbook
import pytest


# @pytest.mark.dataHandling
def test_jsonFileHandling():
    with open("testData\\creds.json") as data:
        finalData = json.load(data)
        print(finalData)
        finalData["negCreds"]["username"]

# @pytest.mark.dataHandling
def test_csvFileHandling():
    with open("testData\\credentails.csv") as data:
        finalData = csv.DictReader(data)
        listData = list(finalData)
        print(listData[-1]["username"])



# @pytest.mark.dataHandling
def test_csvFileHandlingWriting():
    with open("testData\\credentails.csv", mode="w", newline="") as data:
        finalData = csv.DictWriter(data, fieldnames=["username","password" ])
        finalData.writeheader()
        dataV = [{'username': 'tripur123_7', 'password': 'welcome123'}, {'username': 'tripur123_8', 'password': 'welcome123'}]
        for i in dataV:
            finalData.writerow(i)

#pip install openpyxl
# @pytest.mark.dataHandling
def test_excelhandling():
    excelData = load_workbook("testData/sample_creds.xlsx")
    sheet = excelData["sheet1"]
    sheet2 = excelData["Sheet2"]
    values = []
    for i in sheet.iter_rows(min_row=2, values_only=True):
        values.append(i)

    for i in sheet2.iter_rows(min_row=2, values_only=True):
        values.append(i)


    print(values)



def test_excelhandling_write():
    excelData = load_workbook("testData/sample_creds.xlsx")
    sheet = excelData["sheet1"]
    # sheet["A3"]="testing_990"
    sheet.delete_rows(3, sheet.max_row)
    sheet.append(["testing_1", "testing_1"])
    # sheet.insert_rows([["testing_1", "testing_1"],["testing_1", "testing_1"]])
    


    excelData.save("testData/sample_creds.xlsx")



# @pytest.mark.dataHandling
def test_cli():
    data_username = os.getenv("data_usn")
    data_pw = os.getenv("data_password")

    print(data_username)
    print(data_pw)


#pip install python-dotenv
# @pytest.mark.dataHandling
def test_cli():
    load_dotenv(os.getenv('ENV_PATH'))
    data_username1 = os.getenv("data_usn1")
    data_pw1 = os.getenv("data_password1")
    data_url1 = os.getenv("url1")

    print(data_username1)
    print(data_pw1)
    print(data_url1)



@pytest.mark.dataHandling
@pytest.mark.parametrize("a,b,c", [(2,3,5), (3,4,7)])
def test_sum(a,b,c):
    print(a,b,c)
    # assert a + b == c

# def test_sum():
#     a = 20
#     b = 30
#     assert a + b == 50  

       
       




